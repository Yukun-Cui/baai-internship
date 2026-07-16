#!/usr/bin/env python3
"""Extract failed Moore Threads (摩尔线程 / Mthreads) operator names from the Excel results spreadsheet.

This script reads the national-GPU operator test spreadsheet and outputs the list of
operators that failed accuracy tests or were skipped on the Moore Threads (摩尔线程)
backend, suitable for use with `orchestrator.py --mthreads`.

The Moore Threads result column is auto-detected by matching header text containing
"摩尔线程" (and optionally "结果"/"测试结果"). Override with --result-col if needed.
"""

import argparse
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Error: 'openpyxl' is required. Install with: pip install openpyxl")
    sys.exit(1)


def _norm(s) -> str:
    """Normalize a header/cell string: drop newlines and surrounding whitespace."""
    if s is None:
        return ""
    return str(s).replace("\n", "").strip()


def extract_failed_ops(excel_path: str, output_path: str, sheet_name: str = None,
                       result_col: int = None):
    """Extract Moore Threads-failed operator names from Excel spreadsheet.

    An operator is considered "failed" if its Moore Threads result column is:
    - "失败" (accuracy failure), or
    - "跳过" (skipped), or
    - an unrecognized non-numeric, non-pass value.

    Numeric values (e.g. benchmark speedup) and "成功"/"通过" count as pass.

    Returns the sorted list of unique operator names.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Select sheet: explicit name, else first sheet containing operator data.
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = None
        for sname in wb.sheetnames:
            sheet = wb[sname]
            for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
                if any(cell and isinstance(cell, str) and "算子" in cell for cell in row):
                    ws = sheet
                    break
            if ws:
                break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

    print(f"Using sheet: {ws.title}")

    # Find column indices from header row.
    headers = [cell.value for cell in ws[1]]
    op_col = None
    detected_result_col = None

    for i, h in enumerate(headers):
        h_str = _norm(h)
        if not h_str:
            continue
        if "算子" in h_str and "名称" in h_str:
            op_col = i
        elif "摩尔线程" in h_str:
            # Prefer a column that also mentions a result, else take the first 摩尔线程 column.
            if detected_result_col is None or "结果" in h_str:
                detected_result_col = i
        elif op_col is None and ("算子" in h_str or h_str.lower() in ("op", "operator")):
            op_col = i

    if op_col is None:
        print("Error: Could not find operator name column in spreadsheet header")
        print(f"Headers found: {headers}")
        sys.exit(1)

    if result_col is None:
        result_col = detected_result_col
    if result_col is None:
        print("Error: Could not find Moore Threads (摩尔线程) result column in header.")
        print(f"Headers found: {headers}")
        print("Hint: pass --result-col <0-based-index> explicitly.")
        sys.exit(1)

    print(f"Columns found: op_col={op_col}, result_col={result_col}")

    def is_numeric(s: str) -> bool:
        try:
            float(s)
            return True
        except (ValueError, TypeError):
            return False

    failed_ops = set()
    skipped_count = 0
    accuracy_fail_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        op_name = row[op_col] if op_col < len(row) else None
        op_name = _norm(op_name)
        if not op_name or op_name == "None":
            continue

        # Normalize: strip aten:: prefix and overload suffix.
        if op_name.startswith("aten::"):
            op_name = op_name[len("aten::"):]
        if "." in op_name:
            op_name = op_name.split(".")[0]

        result_val = _norm(row[result_col]) if result_col < len(row) else ""

        is_failed = False
        reason = ""
        if result_val == "失败":
            is_failed, reason = True, "accuracy_fail"
            accuracy_fail_count += 1
        elif result_val == "跳过":
            is_failed, reason = True, "skipped"
            skipped_count += 1
        elif result_val in ("成功", "通过", "", "None") or is_numeric(result_val):
            pass  # passed / empty / numeric benchmark value
        else:
            is_failed, reason = True, f"unknown_status({result_val})"
            accuracy_fail_count += 1

        if is_failed:
            failed_ops.add(op_name)
            print(f"  [{reason}] Row {row_idx}: {op_name}")

    wb.close()

    sorted_ops = sorted(failed_ops)
    with open(output_path, "w") as f:
        f.write("\n".join(sorted_ops) + "\n")

    print(f"\nTotal failed operators extracted: {len(sorted_ops)}")
    print(f"  Accuracy failures: {accuracy_fail_count}")
    print(f"  Skipped: {skipped_count}")
    print(f"Output written to: {output_path}")

    return sorted_ops


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    default_excel = os.path.join(
        os.path.dirname(script_dir),
        "operator_test_statistics",
        "第一批及格算子国产GPU测试.xlsx",
    )
    default_output = os.path.join(script_dir, "ops_list_mthreads.txt")

    parser = argparse.ArgumentParser(
        description="Extract failed Moore Threads (摩尔线程) operators from Excel spreadsheet"
    )
    parser.add_argument("-i", "--input", default=default_excel,
                        help=f"Path to Excel spreadsheet (default: {default_excel})")
    parser.add_argument("-o", "--output", default=default_output,
                        help=f"Output path for operator list (default: {default_output})")
    parser.add_argument("-s", "--sheet", default=None,
                        help="Sheet name to use (default: auto-detect)")
    parser.add_argument("--result-col", type=int, default=None,
                        help="0-based index of the Moore Threads result column (default: auto-detect)")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"Error: Excel file not found: {args.input}")
        sys.exit(1)

    extract_failed_ops(args.input, args.output, args.sheet, args.result_col)


if __name__ == "__main__":
    main()
